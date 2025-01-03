import discord
import openpyxl

TOKEN = open("token", "r").readline()
XLSX = "heart.xlsx"


intents = discord.Intents.default()
intents.message_content = True

client = discord.Client(intents=intents)

mention = "<@1318786068795101254>"

@client.event # 데코레이터 - 이벤트 등록
async def on_ready(): # 봇이 로깅을 끝내고 여러가지를 준비한 뒤 호출
    print(f'We have logged in as {client.user}')

@client.event
async def on_message(message):
    if message.content.startswith(f"{mention}"):
        id = str(message.author.id)

        heart = 10

        if message.content.endswith("체력"):
            openxl = openpyxl.load_workbook(XLSX)
            wb = openxl.active
            for i in range(2, 100):
                if wb["A" + str(i)].value == "_" or wb["A" + str(i)].value == id:
                    if wb["D" + str(i)].value > 9:
                        wb["B" + str(i)].value = int(wb["B" + str(i)].value) + int(heart)
                        wb["D" + str(i)].value = int(wb["D" + str(i)].value) - int(heart)
                        await message.channel.send(f"체력 상승 + {heart}", reference=message)
                    else:
                        await message.channel.send("잔여 스탯이 부족합니다.", reference=message)

                    break
            openxl.save(XLSX)

        elif message.content.endswith("공격력"):
            openxl = openpyxl.load_workbook(XLSX)
            wb = openxl.active
            for i in range(2, 100):
                if wb["A" + str(i)].value == "_" or wb["A" + str(i)].value == id:
                    if wb["D" + str(i)].value > 9:
                        wb["C" + str(i)].value = int(wb["C" + str(i)].value) + int(heart)
                        wb["D" + str(i)].value = int(wb["D" + str(i)].value) - int(heart)
                        await message.channel.send(f"공격력 상승 + {heart}", reference=message)
                    else:
                        await message.channel.send("잔여 스탯이 부족합니다.", reference=message)

                    break
            openxl.save(XLSX)

        elif message.content.endswith("등록"):
            openxl = openpyxl.load_workbook(XLSX)
            wb = openxl.active
            for i in range(2, 100):
                if wb["A" + str(i)].value == "_" or wb["A" + str(i)].value == id:
                    if wb["A" + str(i)].value == id:
                        await message.channel.send(f"이미 등록된 사용자입니다.", reference=message)
                    else:
                        wb["A" + str(i)].value = id
                        wb["E" + str(i)].value = str(message.author.display_name)
                        wb["D" + str(i)].value = 100
                        await message.channel.send(f"정보가 등록 되었습니다.", reference=message)

                    break
            openxl.save(XLSX)

    if message.content.startswith("?체력"):
        openxl = openpyxl.load_workbook(XLSX)
        wb = openxl.active
        id = str(message.author.id)
        for i in range(2, 100):
            if wb["A" + str(i)].value == "_" or wb["A" + str(i)].value == id:
                if wb["A" + str(i)].value == id:
                    await message.channel.send(f"{message.author.display_name} 님의 체력은 {wb['B' + str(i)].value} 입니다.", reference=message)
                else:
                    await message.channel.send(f"{message.author.display_name} 님의 체력은 0입니다.", reference=message)
                break

    elif message.content.startswith("?공격력"):
        openxl = openpyxl.load_workbook(XLSX)
        wb = openxl.active
        id = str(message.author.id)
        for i in range(2, 100):
            if wb["A" + str(i)].value == "_" or wb["A" + str(i)].value == id:
                if wb["A" + str(i)].value == id:
                    await message.channel.send(f"{message.author.display_name} 님의 공격력은 {wb['C' + str(i)].value} 입니다.", reference=message)
                    break
                else:
                    await message.channel.send(f"{message.author.display_name} 님의 공격력은 0입니다.", reference=message)
                    break

    elif message.content.startswith("?스탯"):
        openxl = openpyxl.load_workbook(XLSX)
        wb = openxl.active
        id = str(message.author.id)
        for i in range(2, 100):
            if wb["A" + str(i)].value == "_" or wb["A" + str(i)].value == id:
                if wb["A" + str(i)].value == id:
                    await message.channel.send(f"{message.author.display_name} 님의 잔여 스탯은 {wb['D' + str(i)].value} 입니다.", reference=message)
                    break
                else:
                    await message.channel.send(f"{message.author.display_name} 님의 잔여 스탯은 0입니다.", reference=message)
                    break

    elif message.content.startswith("/잠뜰TV 보실래요?"):
        openxl = openpyxl.load_workbook(XLSX)
        wb = openxl.active
        for i in range(2, 100):
            wb["A" + str(i)].value = "_"
            wb["B" + str(i)].value = 0
            wb["C" + str(i)].value = 0
            wb["D" + str(i)].value = ""
            wb["E" + str(i)].value = ""
        await message.channel.send("초기화 완료.")
        openxl.save(XLSX)

client.run(TOKEN)